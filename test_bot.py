import logging
from telegram import ReplyKeyboardMarkup
from telegram.ext import Updater, CommandHandler, ConversationHandler, MessageHandler, Filters
import openpyxl
import settings
import re
from bs4 import BeautifulSoup
from urllib.request import Request, urlopen
import urllib.request
import config
import glob
from urllib.parse import urlparse
import os.path
import schedule
import time
from threading import Thread

logging.basicConfig(filename="bot.log", level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


# def schedule_checker():
#     while True:
#         schedule.run_pending()
#         time.sleep(1)


def working_with_files(dp):
    dp.add_handler(MessageHandler("fffffffffff"))
    print("I'm working...")
    DATASET_URL = "https://ptgh.onego.ru/9006/"
    url = Request(DATASET_URL)
    html_page = urlopen(url)
    soup = BeautifulSoup(html_page, "html.parser")
    link = []
    for item in soup.findAll('a'):
        link.append(item.get('href'))

    for index, elem in enumerate(link):
        if (elem.find('ismen_nov')) != -1:
            new_elem = urlparse(elem)
            print(new_elem[2].replace('/', '_'))
            break

    file_list = glob.glob('*.xlsx')
    print(file_list)
    if not file_list:
        print('excel файлов нет')
    else:
        if file_list[0] == new_elem[2].replace('/', '_'):
            print('нового расписания нет')
            pass
        else:
            file_xlsx = urllib.request.urlopen(elem).read()
            f = open(new_elem[2].replace('/', '_'), "wb")
            f.write(file_xlsx)
            f.close()
            print('доступно новое расписание')
            dp.add_handler(MessageHandler("fffffffffff"))
    file_list = glob.glob('*.xlsx')
    for item in file_list:
        if item != new_elem[2].replace('/', '_'):
            os.remove(item)


def greet_user(update, context):
    print("Вызван /start")
    my_keyboard = ReplyKeyboardMarkup([['Расписание']], resize_keyboard=True)
    update.message.reply_text("Привет, чтобы узнать расписание нажми кнопку", reply_markup=my_keyboard)


def dialog_start(update, context):
    book = openpyxl.open(glob.glob('*ismen*.xlsx'), read_only=True)
    my_keyboard = ReplyKeyboardMarkup([book.sheetnames], resize_keyboard=True)
    update.message.reply_text("Изменения к расписанию занятий Корпус 1 (ул. Мурманская, д. 30) \n выберите дату:", reply_markup=my_keyboard)
    return "step_one"


def choose_sheet(update, context):
    book = openpyxl.open(config.ISMEN_FILE_NAME, read_only=True)
    user_text = update.message.text
    context.user_data["dialog"] = {"sheet": user_text}
    sheet = book[context.user_data["dialog"]["sheet"]]
    schedules = []
    for row in range(1, sheet.max_row):
        a_column = sheet[row][0].value
        if a_column is None:
            a_column = ''
        c_column = sheet[row][2].value
        if c_column is None or c_column == 'ауд.':
            c_column = ''
        b_column = sheet[row][1].value
        if b_column is None:
            continue
        else:
            b_column_split = ' '.join(b_column.split())
            result_abc = str(a_column) + str(b_column_split) + str(c_column)
            schedules.append(result_abc)
    for row in range(1, sheet.max_row):
        e_column = sheet[row][4].value
        if e_column is None:
            e_column = ''
        g_column = sheet[row][6].value
        if g_column is None or g_column == 'ауд.':
            g_column = ''
        f_column = sheet[row][5].value
        if f_column is None:
            continue
        else:
            f_column_split = ' '.join(f_column.split())
            result_efg = str(e_column) + str(f_column_split) + str(g_column)
            schedules.append(result_efg)
    groups = re.compile(r'([ЭМТВ]\s[0-4][0-4])|([ЭМТВ]\s[0-4])|([БУИП][ДС]\s[0-4][0-4]|[ИБПУ][СД]\s[0-4])|([П][С][О]\s[0-4][0-4])|([ВМ][0-4]|[У][Д][0-4])')
    idx1 = [i for i, item in enumerate(schedules) if re.search(groups, item)]
    dict_groups = {}
    for i in range(len(idx1)-1):
        slices = schedules[idx1[i]:idx1[i+1]]
        dict_groups[slices[0]] = slices[1:]
    last_slice = schedules[idx1[-1]:]
    dict_groups[last_slice[0]] = last_slice[1:]
    groups_list = list(dict_groups.keys())
    n = 4
    group_names = [groups_list[i * n:(i + 1) * n] for i in range((len(groups_list) + n - 1) // n)]
    my_keyboard = ReplyKeyboardMarkup(group_names, resize_keyboard=True)
    update.message.reply_text("Выберите группу:", reply_markup=my_keyboard)
    return "step_two"


def print_schedules(update, context):
    context.user_data["dialog"]["group"] = update.message.text
    book = openpyxl.open(config.ISMEN_FILE_NAME, read_only=True)
    sheet = book[context.user_data["dialog"]["sheet"]]
    schedules = []
    for row in range(1, sheet.max_row):
        a_column = sheet[row][0].value
        if a_column is None:
            a_column = ''
        c_column = sheet[row][2].value
        if c_column is None or c_column == 'ауд.':
            c_column = ''
        b_column = sheet[row][1].value
        if b_column is None:
            continue
        else:
            b_column_split = ' '.join(b_column.split())
            result_abc = str(a_column) + str(b_column_split) + str(c_column)
            schedules.append(result_abc)
    for row in range(1, sheet.max_row):
        e_column = sheet[row][4].value
        if e_column is None:
            e_column = ''
        g_column = sheet[row][6].value
        if g_column is None or g_column == 'ауд.':
            g_column = ''
        f_column = sheet[row][5].value
        if f_column is None:
            continue
        else:
            f_column_split = ' '.join(f_column.split())
            result_efg = str(e_column) + str(f_column_split) + str(g_column)
            schedules.append(result_efg)
    groups = re.compile(r'([ЭМТВ]\s[0-4][0-4])|([ЭМТВ]\s[0-4])|([БУИП][ДС]\s[0-4][0-4]|[ИБПУ][СД]\s[0-4])|([П][С][О]\s[0-4][0-4])|([ВМ][0-4]|[У][Д][0-4])')
    idx1 = [i for i, item in enumerate(schedules) if re.search(groups, item)]
    dict_groups = {}
    for i in range(len(idx1)-1):
        slices = schedules[idx1[i]:idx1[i+1]]
        dict_groups[slices[0]] = slices[1:]
    last_slice = schedules[idx1[-1]:]
    dict_groups[last_slice[0]] = last_slice[1:]
    selected_group = str(dict_groups[context.user_data["dialog"]["group"]]).replace(',','\n')
    my_keyboard = ReplyKeyboardMarkup([['Расписание']], resize_keyboard=True)
    update.message.reply_text(selected_group, reply_markup=my_keyboard)
    return ConversationHandler.END


def main():
    
    # Thread(target=schedule_checker).start()
    mybot = Updater(settings.API_KEY, use_context=True)
    dp = mybot.dispatcher
    dialog = ConversationHandler(
        entry_points=[
            MessageHandler(Filters.regex('^([Рр]асписание)$'), dialog_start)
        ],
        states={
            "step_one": [MessageHandler(Filters.regex('^((3[01]|[12][0-9]|0[1-9]).(1[0-2]|0[1-9]))$'), choose_sheet)],
            "step_two": [MessageHandler(Filters.regex('(([ЭМТВ].[0-4][0-4])|([ЭМТВ].[0-4])|([БУИП][ДС].[0-4][0-4]|[ИБПУ][СД].[0-4])|([П][С][О].[0-4][0-4])|([ВМ][0-4]|[У][Д][0-4]))'), print_schedules)]
        },
        fallbacks=[]
    )
    schedule.every(10).seconds.do(working_with_files, dp)
    while True:
        schedule.run_pending()
        time.sleep(1)
    dp.add_handler(dialog)
    dp.add_handler(CommandHandler("start", greet_user))
    # dp.add_handler(MessageHandler('привет'))
    logging.info("bot started")
    mybot.start_polling()
    mybot.idle()
    


if __name__ == "__main__":
    main()


