import os
import openpyxl
wb = openpyxl.load_workbook(os.path.join('ismen_nov.xlsx'))
ws = wb['22.12']

print(wb.sheetnames)
print(ws.title)


def parsing():
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            else:
                print(cell.value, end=' ')
        print()


parsing()
