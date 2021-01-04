import openpyxl
book = openpyxl.open('ismen_nov.xlsx', read_only=True)


def b_column():
    for row in range(1, sheet.max_row):
        a_column = sheet[row][0].value
        if a_column is None:
            a_column = ''
        c_column = sheet[row][2].value
        if c_column is None:
            c_column = ''
        b_column = sheet[row][1].value
        if b_column is None:
            continue
        else:
            b_column_split = ' '.join(b_column.split())
            result_abc = str(a_column) + str(b_column_split) + ' ' + str(c_column)
            print(result_abc)


def f_column():
    for row in range(1, sheet.max_row):
        e_column = sheet[row][4].value
        if e_column is None:
            e_column = ''
        g_column = sheet[row][6].value
        if g_column is None:
            g_column = ''
        f_column = sheet[row][5].value
        if f_column is None:
            continue
        else:
            f_column_split = ' '.join(f_column.split())
            result_efg = str(e_column) + str(f_column_split) + ' ' + str(g_column)
            print(result_efg)


print(book.sheetnames)
sheet_name = input('Выберите день: \n') 
sheet = book[sheet_name]

choise = input('Нажмите 0 для групп: УД31, Т31, Э12, БД12, ПСО11, ПСО21, ПСО31, ПД13, ПД23, ПД32, УД01, УД21, В21, Т21, ИС11\nНажмите 1 для групп: Э22, БД22, ПСО12, ПСО22, ПД12, ПД22, ПД24, ПД33, УД11, В01, Т11, М01, ИС21\n')
if choise == "0":
    print(b_column())
else:
    print(f_column())
