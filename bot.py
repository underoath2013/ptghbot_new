import time
from urllib.error import URLError

from bs4 import BeautifulSoup
from db import db, get_or_create_user, subscribe_user, unsubscribe_user, \
    get_subsribed
import glob
import logging

import openpyxl
import os.path
import system_functions
from random import choice
import pickle
import re
import requests
from requests import RequestException
import settings
import sys
from telegram import ReplyKeyboardMarkup
from telegram.ext import Updater, CommandHandler, ConversationHandler, \
    MessageHandler, Filters
from telegram.error import BadRequest, NetworkError, Unauthorized
from urllib.request import Request, urlopen
from urllib.parse import urlparse

logging.basicConfig(
    filename="bot.log", level=logging.INFO, format=
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)


# chat_id = update.effective_chat.id - получить chat.id текущего пользователя


def bot_commands(update, _):
    """ Доступные команды для работы с ботом """
    update.message.reply_text("/start - старт бота\n"
                              "Проверяю наличие нового расписания"
                              " каждые 2 часа\n"
                              "/subscribe - получать уведомление"
                              " о новом расписании\n"
                              "/unsubscribe - отписаться от уведомлений\n"
                              "автор бота @drbrch,"
                              " принимаю в лс предложения/замечания")


def main_keyboard():
    """ Клавиатура главного меню """
    return ReplyKeyboardMarkup(
        [['Скачать основное', 'Скачать изменения'],
         ['Просмотреть изменения'],
         ['Команды бота', 'Звонки']], resize_keyboard=True
    )


def show_dontunderstand(update, _):
    update.message.reply_text("Я вас не понял, лучше нажмите кнопку")


def cancel_dialog(update, _):
    update.message.reply_text('Жаль', reply_markup=main_keyboard())
    return ConversationHandler.END


def subscribe(update, _):
    """ При вызове этой функции пользователь подписывается на рассылку
    уведомлений """
    user = get_or_create_user(db, update.effective_user, update.message.chat.id)
    subscribe_user(db, user)
    update.message.reply_text('Уведомления о новом расписании подключены')


def unsubscribe(update, _):
    """ При вызове этой функции пользователь отписывается от рассылки
        уведомлений """
    user = get_or_create_user(db, update.effective_user, update.message.chat.id)
    unsubscribe_user(db, user)
    update.message.reply_text('Уведомления о новом расписании отключены')


def greet_user(update, _):
    """ Приветствует пользователя, выводит основную клавиатуру """
    user = get_or_create_user(db, update.effective_user, update.message.chat.id)
    print("Вызван /start")
    update.message.reply_text("Привет, чтобы узнать расписание 1 корпуса\n"
                              "нажми нужную тебе кнопку\n"
                              "Уведомления об обновлении расписания\n"
                              "можно подключить в 'Команды бота'",
                              reply_markup=main_keyboard())


def show_rings(update, context):
    """ Выводит список звонков. Также будет выводить в случайном порядке
     любые jpeg файлы в каталоге images """
    user = get_or_create_user(db, update.effective_user, update.message.chat.id)
    rings_list = glob.glob('images/rings*.jp*g')
    # rings_pic_filename = choice(rings_list) отправка случайного jpeg
    for picture in rings_list:
        context.bot.send_photo(chat_id=user['chat_id'],
                               photo=open(picture, 'rb'),
                               reply_markup=main_keyboard())


def sending_notify_about_updating_schedules(schedule_folder, context):
    """ Присылает уведомления, подписанным на уведомления пользоватеям """
    for user in get_subsribed(db):
        try:
            context.bot.send_message(chat_id=user['chat_id'],
                                     text=f'{schedule_folder} обновлено')
        except Unauthorized:  # ловим ошибку, если юзер заблокировал бота
            # выводим в логи юзеров, заблокировавших бота
            logging.info(f"User {user['chat_id']} has blocked the bot")


def parsing_links_from_schedule_html_downloading_schedules(context, user=False):
    """ Парсит страницу DATASET_URL, ищет ссылку на Изменения к основному
     расписанию, вызывает функцию скачивания downloading_schedules, передавая
     в нее параметры, необходимые для скачивания Изменений.
     Ищет ссылку на Основное расписание, вызывает функцию скачивания
     downloading_schedules, передавая в нее параметры, необходимые для
     скачивания Основного """
    print('*** Start parsing ***')
    global NAME_OF_CHANGES_SCHEDULE_FILE
    NAME_OF_CHANGES_SCHEDULE_FILE = ''
    global NAME_OF_MAIN_SCHEDULE_FILE
    NAME_OF_MAIN_SCHEDULE_FILE = ''
    DATASET_URL = "https://ptgh.onego.ru/9006/"
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0)'
                      ' Gecko/20100101 Firefox/91.0'
    }
    try:
        html_page = requests.get(DATASET_URL, headers=headers, timeout=2)
    except RequestException:
        logging.error('Site unavailable')
    else:
        soup = BeautifulSoup(html_page.text, "html.parser")
        links = []
        # нет селектора, поиск идет по всем тегам <a>
        for item in soup.findAll('a'):
            links.append(item.get('href'))
        for index, link in enumerate(links):  # кол-во итераций = кол-ву ссылок
            if (link.find('ismen_nov')) != -1:
                changes_schedule_link = link
                changes_schedule = urlparse(changes_schedule_link)
                NAME_OF_CHANGES_SCHEDULE_FILE = changes_schedule.path.replace(
                    '/', '')
                print(NAME_OF_CHANGES_SCHEDULE_FILE)
                # передаем ссылку по которой качать, название папки,
                # имя скачиваемого файла, context
                download_result = downloading_schedules(
                    changes_schedule_link, 'changes_schedule',
                    NAME_OF_CHANGES_SCHEDULE_FILE)
                if download_result is True:
                    # если есть новое расписание, то парсим его полностью и формируем pickle файл
                    parsing_data(NAME_OF_CHANGES_SCHEDULE_FILE)
                    if user is False:
                    # дополнительно, если отработал job_queue, то отсылаем уведомление юзеру
                        sending_notify_about_updating_schedules(
                            'changes_schedule', context)
                    continue  # переход к следующей итерации
            if re.search(r'РАСПИСАНИЕ.*\.xlsx', link):
                main_schedule_link = link
                main_schedule = urlparse(main_schedule_link)
                NAME_OF_MAIN_SCHEDULE_FILE = main_schedule.path.replace('/', '')
                print(NAME_OF_MAIN_SCHEDULE_FILE)
                # передаем ссылку по которой качать, название папки,
                # имя скачиваемого файла, context
                download_result = downloading_schedules(
                    main_schedule_link, 'main_schedule',
                    NAME_OF_MAIN_SCHEDULE_FILE)
                if download_result is True and user is False:
                    sending_notify_about_updating_schedules(
                        'main_schedule', context)


def downloading_schedules(
        schedule_link, schedule_folder, name_of_schedule_file):
    """ Скачивает Изменения и Основное расписание
    :param schedule_link: ссылка для скачивания файла
    :type schedule_link: str
    :param schedule_folder: директория в которой находится файл
    :type schedule_folder: str
    :param name_of_schedule_file: имя файла с расписанием
    :type name_of_schedule_file: str
    """
    if len(os.listdir(schedule_folder)) == 0:  # listdir возвращает список
        schedule_xlsx = requests.get(schedule_link)
        f = open(schedule_folder + '/' + name_of_schedule_file, "wb")
        f.write(schedule_xlsx.content)  # записывает content из get запроса
        f.close()
        print(f'Расписание скачано в {schedule_folder}')
        return True  # запомним результат как True, если функция отработала
    elif name_of_schedule_file not in os.listdir(schedule_folder):
        for file in os.listdir(schedule_folder):
            os.remove(schedule_folder + '/' + file)
        schedule_xlsx = requests.get(schedule_link)
        f = open(schedule_folder + '/' + name_of_schedule_file, "wb")
        f.write(schedule_xlsx.content)  # записывает content из get запроса
        f.close()
        print(f'Расписание обновлено в {schedule_folder}')
        return True  # запомним результат как True, если функция отработала
    return False  # запомним результат как False, если функция не отработала


def parsing_data(name_of_schedule_file):
    """ Парсит данные их эксель файла в структуру словаря, формирует pickle dump
        :param name_of_schedule_file: название файла с расписанием
        :type name_of_schedule_file: str
        """
    book = openpyxl.open(
        'changes_schedule/' + name_of_schedule_file,
        read_only=True
    )
    pickled_schedule_dict = {}
    for sheetname in book.sheetnames:
        parsed_schedule = system_functions.parsing_changes_xlsx(book[sheetname])
        groups = re.compile(
            r'(БД 12)|(БД 22)|(ИС 11)|(ИС 21)|(ИС 31)|'
            r'(В 01)|(В 11)|(В 21)|(В 31)|'
            r'(ЗИ 11)|'
            r'(М 01)|(М 11)|(М 21)|(М 31)|'
            r'(ПД 12)|(ПД 13)|(ПД 22)|(ПД 23)|(ПД 24)|(ПД 32)|(ПД 33)|(ПД 34)|'
            r'(ПСО 11)|(ПСО 12)|(ПСО 21)|(ПСО 22)|(ПСО 31)|(ПСО 32)|'
            r'(Т 11)|(Т 21)|(Т 31)|'
            r'(УД 01)|(УД 11)|(УД 21)|(УД 31)|(Э 12)|(Э 22)')
        idx1 = [i for i, item in enumerate(parsed_schedule) if
                re.search(groups, item)]
        dict_groups = {}
        for i in range(len(idx1) - 1):
            slices = parsed_schedule[idx1[i]:idx1[i + 1]]
            dict_groups[slices[0]] = slices[1:]
        try:
            last_slice = parsed_schedule[idx1[-1]:] # exception here
            dict_groups[last_slice[0]] = last_slice[1:]
            pickled_schedule_dict[sheetname] = dict_groups
        except IndexError:
            pickled_schedule_dict[sheetname] = 'На этот день еще ничего нет'
    with open('changes_schedule/pickled_schedule_dict.pickle', 'wb') as f:
        pickle.dump(pickled_schedule_dict, f)


def send_main_schedule(update, context):
    parsing_links_from_schedule_html_downloading_schedules(context, user=True)
    user = get_or_create_user(db, update.effective_user, update.message.chat.id)
    if NAME_OF_MAIN_SCHEDULE_FILE == '':
        update.message.reply_text('Сайт недоступен или основное расписание '
                                  'не найдено на сайте, '
                                  'но у меня есть старая версия, '
                                  'используй внимательно')
        context.bot.send_document(chat_id=user['chat_id'], document=open(
            glob.glob("main_schedule/*.xlsx")[0], 'rb'), reply_markup=main_keyboard())
    else:
        context.bot.send_document(
            chat_id=user['chat_id'], document=open(
                'main_schedule/' + NAME_OF_MAIN_SCHEDULE_FILE, 'rb'),
            reply_markup=main_keyboard()
        )


def send_changes_schedule(update, context):
    parsing_links_from_schedule_html_downloading_schedules(context, user=True)
    user = get_or_create_user(db, update.effective_user, update.message.chat.id)
    if NAME_OF_CHANGES_SCHEDULE_FILE == '':
        update.message.reply_text('Сайт недоступен или изменений расписания '
                                  'не найдено на сайте, '
                                  'но у меня есть старая версия, '
                                  'используй внимательно')
        context.bot.send_document(chat_id=user['chat_id'], document=open(
            glob.glob("changes_schedule/*.xlsx")[0], 'rb'), reply_markup=main_keyboard())
    else:
        context.bot.send_document(
            chat_id=user['chat_id'], document=open(
                'changes_schedule/' + NAME_OF_CHANGES_SCHEDULE_FILE,
                'rb'), reply_markup=main_keyboard())


def show_dates_of_changes_schedule(update, _):
    """ Выводит список дат, доступных для выбора """
    if NAME_OF_CHANGES_SCHEDULE_FILE == '':
        update.message.reply_text('Изменений расписания не найдено на сайте')
    else:
        with open('changes_schedule/' + 'pickled_schedule_dict.pickle', 'rb') as f:
            pickled_schedule_dict = pickle.load(f)
        dates_list = list(pickled_schedule_dict.keys())
        n = 6
        dates_names = [dates_list[i * n:(i + 1) * n] for i in range(
            (len(dates_list) + n - 1) // n)]
        # добавляем кнопку отмена к кнопкам с датами
        dates_names.append(['Отмена'])
        # старый вариант клавиатуры
        # my_keyboard = ReplyKeyboardMarkup(
        #     # выводим кнопки с датами и кнопку отмена
        #     [pickled_schedule_dict.keys(), ['Отмена']], resize_keyboard=True)
        my_keyboard = ReplyKeyboardMarkup(dates_names, resize_keyboard=True)
        update.message.reply_text(
            "Изменения к расписанию занятий Корпус 1 (ул. Мурманская, д. 30)\n"
            "выберите дату:",
            reply_markup=my_keyboard)
        return "step_one"


def choose_sheet_of_changes_schedule(update, context):
    """ Парсит по именам группы расписание на выбранную пользователем дату,
    формирует словарь dict_groups и записывает его во встроенный словарь
    context.user_data, выводит списко групп пользователю """
    user_text = update.message.text
    context.user_data["dialog"] = {"day": user_text}
    with open('changes_schedule/' + 'pickled_schedule_dict.pickle', 'rb') as f:
        pickled_schedule_dict = pickle.load(f)
    schedule_for_selected_day = pickled_schedule_dict[context.user_data["dialog"]["day"]]
    #context.user_data["dict_groups"] = dict_groups ранее сохраняли во встроенный словарь user data
    if isinstance(schedule_for_selected_day, str):
        update.message.reply_text(schedule_for_selected_day, reply_markup=main_keyboard())
        return ConversationHandler.END
    groups_list = list(schedule_for_selected_day.keys())
    n = 4
    groups_names = [groups_list[i * n:(i + 1) * n] for i in range(
        (len(groups_list) + n - 1) // n)]
    # добавляем кнопку отмена к кнопкам с названием групп
    groups_names.append(['Отмена'])
    my_keyboard = ReplyKeyboardMarkup(
        groups_names, resize_keyboard=True)
    update.message.reply_text("Выберите группу:", reply_markup=my_keyboard)
    return "step_two"


def print_changes_schedule(update, context):
    """ Печатает расписание для выбранной пользователем группы """
    with open('changes_schedule/' + 'pickled_schedule_dict.pickle', 'rb') as f:
        pickled_schedule_dict = pickle.load(f)
    context.user_data["dialog"]["group"] = update.message.text
    #context.user_data["dict_groups"][context.user_data["dialog"]["group"]] ранее читали из встроенного словаря user data
    # context.user_data.clear() позволяет очищать словарь context.user_data
    schedule_of_selected_group = pickled_schedule_dict[context.user_data["dialog"]["day"]][context.user_data["dialog"]["group"]]
    if len(schedule_of_selected_group) > 0:
        update.message.reply_text(
            str(schedule_of_selected_group).replace(
                ', ', '\n').replace('[', '').replace(']', '').replace("'", ""),
            reply_markup=main_keyboard())
    else:
        update.message.reply_text("Удачи!", reply_markup=main_keyboard())
    return ConversationHandler.END


# начало блока чтения и парсинга основного расписания, сейчас недоступен, т.к. основное выложено полностью за все дни
# из def main_keyboard() убрана кнопка просмотра основного расписания
def show_dates_of_main_schedule(update, _):
    """ Выводит список дат, доступных для выбора основного расписания"""
    if NAME_OF_MAIN_SCHEDULE_FILE == '':
        update.message.reply_text('Основного расписания не найдено на сайте')
    else:
        book = openpyxl.open(
            'main_schedule/' + NAME_OF_MAIN_SCHEDULE_FILE,
            read_only=True
        )
        my_keyboard = ReplyKeyboardMarkup(
            # выводим кнопки с датами и кнопку отмена
            [book.sheetnames, ['Отмена']], resize_keyboard=True)
        print(book.sheetnames)
        update.message.reply_text(
            "Основное расписание занятий Корпус 1 (ул. Мурманская, д. 30)\n"
            "выберите дату:",
            reply_markup=my_keyboard)
        return "main_step_one"


def choose_sheet_of_main_schedule(update, context):
    """ Парсит по именам группы основное расписание на выбранную пользователем
     дату, формирует словарь main_schedule_dict и записывает его во встроенный
      словарь context.user_data, выводит список групп пользователю """
    book = openpyxl.open(
        'main_schedule/' + NAME_OF_MAIN_SCHEDULE_FILE,
        read_only=True
    )
    user_text = str(update.message.text)
    print(user_text)
    print(type(user_text))
    context.user_data["dialog"] = {"sheet": user_text}
    sheet = book[context.user_data["dialog"]["sheet"]]
    print(book[context.user_data["dialog"]["sheet"]])
    update.message.reply_text("Загружаю...")
    if sheet[5][2].value is not None:
        parsed_main_schedule = system_functions.parsing_main_xlsx(sheet)
        context.user_data["main_dict_groups"] = parsed_main_schedule
        groups_of_main_list = list(parsed_main_schedule.keys())
        n = 4
        groups_of_main_names = [groups_of_main_list[i * n:(i + 1) * n] for i in range(
            (len(groups_of_main_list) + n - 1) // n)]
        # добавляем кнопку отмена к кнопкам с названием групп
        groups_of_main_names.append(['Отмена'])
        my_keyboard = ReplyKeyboardMarkup(
            groups_of_main_names, resize_keyboard=True)
        update.message.reply_text("Выберите группу:", reply_markup=my_keyboard)
        return "main_step_two"
    else:
        update.message.reply_text("На этот день ничего нет",
                                  reply_markup=main_keyboard())
        return ConversationHandler.END


def print_main_schedule(update, context):
    """ Печатает расписание для выбранной пользователем группы """

    context.user_data["dialog"]["group"] = update.message.text
    main_schedule_of_selected_group = \
        context.user_data["main_dict_groups"][context.user_data["dialog"]["group"]
        ]
    # context.user_data.clear() позволяет очищать словарь context.user_data
    if len(main_schedule_of_selected_group) > 0:
        for index, elem in enumerate(main_schedule_of_selected_group):
            main_schedule_of_selected_group[index] = str(index+1) + ' п. ' + main_schedule_of_selected_group[index]
        main_schedule_of_selected_group = str(main_schedule_of_selected_group).replace(
            ',', '\n').replace('[', '').replace(']', '').replace("'", "")
        # с помощью цикла проходим строку находя /n
        # в этой же итерации убираем пробел за /n, если пробел есть
        update.message.reply_text(main_schedule_of_selected_group,
                                  reply_markup=main_keyboard())
    else:
        update.message.reply_text("Удачи!", reply_markup=main_keyboard())
    return ConversationHandler.END
#конец блока чтения и парсинга основного расписания


def main():
    mybot = Updater(settings.API_KEY, use_context=True)
    jq = mybot.job_queue
    # раз в заданный период и при старте бота запускаем функцию
    # downloading_and_comparing_xlsx_schedules
    jq.run_repeating(parsing_links_from_schedule_html_downloading_schedules,
                     interval=7200, first=1)
    dp = mybot.dispatcher
    # начало диалога с пользователем
    dialog = ConversationHandler(
        entry_points=[
            MessageHandler(Filters.regex('^([Пп]росмотреть изменения)$'),
                           show_dates_of_changes_schedule)
        ],
        states={
            "step_one": [
                MessageHandler(
                    Filters.regex('\d\d.\d\d'),
                    choose_sheet_of_changes_schedule)],
            "step_two": [
                MessageHandler(Filters.regex(
                    r'(БД 12)|(БД 22)|(ИС 11)|(ИС 21)|(ИС 31)|'
                    r'(В 01)|(В 11)|(В 21)|(В 31)|'
                    r'(ЗИ 11)|'
                    r'(М 01)|(М 11)|(М 21)|(М 31)|'
                    r'(ПД 12)|(ПД 13)|(ПД 22)|(ПД 23)|(ПД 24)|'
                    r'(ПД 32)|(ПД 33)|(ПД 34)|'
                    r'(ПСО 11)|(ПСО 12)|(ПСО 21)|(ПСО 22)|(ПСО 31)|(ПСО 32)|'
                    r'(Т 11)|(Т 21)|(Т 31)|'
                    r'(УД 01)|(УД 11)|(УД 21)|(УД 31)|(Э 12)|(Э 22)'),
                    print_changes_schedule)]
        },
        fallbacks=[
            # при нажатии кнопки отмена выходим из диалога
            MessageHandler(Filters.regex('^([Оо]тмена)$'), cancel_dialog),
            MessageHandler(
                # фильтры не дают пользователю ввести лишний текст, прислать
                # фото, видео, геолокацию
                Filters.text | Filters.video | Filters.photo | Filters.document
                | Filters.location, show_dontunderstand)]
    )
    dialog_main = ConversationHandler(
        entry_points=[
            MessageHandler(Filters.regex('^([Пп]росмотреть основное)$'),
                           show_dates_of_main_schedule)
        ],
        states={
            "main_step_one": [
                MessageHandler(
                    Filters.regex('\d\d.\d\d'),
                    choose_sheet_of_main_schedule)],
            "main_step_two": [
                MessageHandler(Filters.regex(
                    r'(БД 12)|(БД 22)|(ИС 11)|(ИС 21)|(ИС 31)|'
                    r'(В 01)|(В 11)|(В 21)|(В 31)|'
                    r'(ЗИ 11)|'
                    r'(М 01)|(М 11)|(М 21)|(М 31)|'
                    r'(ПД 12)|(ПД 13)|(ПД 22)|(ПД 23)|(ПД 24)|'
                    r'(ПД 32)|(ПД 33)|(ПД 34)|'
                    r'(ПСО 11)|(ПСО 12)|(ПСО 21)|(ПСО 22)|(ПСО 31)|(ПСО 32)|'
                    r'(Т 11)|(Т 21)|(Т 31)|'
                    r'(УД 01)|(УД 11)|(УД 21)|(УД 31)|(Э 12)|(Э 22)'),
                    print_main_schedule)]
        },
        fallbacks=[
            # при нажатии кнопки отмена выходим из диалога
            MessageHandler(Filters.regex('^([Оо]тмена)$'), cancel_dialog),
            MessageHandler(
                # фильтры не дают пользователю ввести лишний текст, прислать
                # фото, видео, геолокацию
                Filters.text | Filters.video | Filters.photo | Filters.document
                | Filters.location, show_dontunderstand)]
    )
    dp.add_handler(dialog)
    dp.add_handler(dialog_main)
    dp.add_handler(CommandHandler('subscribe', subscribe))
    dp.add_handler(CommandHandler('unsubscribe', unsubscribe))
    dp.add_handler(CommandHandler('show_rings', show_rings))
    dp.add_handler(CommandHandler("start", greet_user))
    dp.add_handler(MessageHandler(Filters.regex('^([Зз]вонки)$'), show_rings))
    dp.add_handler(MessageHandler(Filters.regex('^([Сс]качать основное)$'),
                                  send_main_schedule))
    dp.add_handler(MessageHandler(Filters.regex('^([Сс]качать изменения)$'),
                                  send_changes_schedule))
    dp.add_handler(
        MessageHandler(Filters.regex('^([Кк]оманды бота)$'), bot_commands))
    logging.info("bot started")
    mybot.start_polling()
    mybot.idle()


if __name__ == "__main__":
    main()
